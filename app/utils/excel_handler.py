import os
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
import datetime
from collections import Counter

from app.utils.convert_decimal_to_binary import convert_decimal_to_binary
from .format_time import format_time
from .get_time_duration import get_time_duration

def create_excel_file(date, server_name):
    month, day, year = date.split("-")
    excel_file_name = f"{year}_{month}_{server_name}_status_flags.xlsx"
    path = os.path.join("./exports/", excel_file_name)
    if not os.path.exists(path):
        source_path = f"./exports/model_{server_name}.xlsx"
        destination_path = path
        shutil.copy(source_path, destination_path)
        os.rename(destination_path, path)

def export_data_into_excel(status_flags, pmu, date, server_name):
    month, day, year = date.split("-")
    if not status_flags:
        return
    try:
        excel_file_name = f"./exports/{year}_{month}_{server_name}_status_flags.xlsx"
        workbook = openpyxl.load_workbook(excel_file_name)
    except FileNotFoundError:
        workbook = Workbook()
    try:
        worksheet = workbook[pmu]
    except KeyError:
        worksheet = workbook.create_sheet(pmu)
    header_in_excel(worksheet, pmu)
    configure_pmu_summary_header(worksheet)
    get_period_time_of_status_flags(worksheet, status_flags, date)
    TOTAL_PERIOD_COLUMN = 12
    total_duration_of_status_flags = get_total_duration_of_status_flags(worksheet)
    worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value = total_duration_of_status_flags
    STATUS_FLAGS_HEX_COLUMN = 3
    most_common_status_flags, amount_of_status_flags  = get_total_items_in_an_excel_column(worksheet, STATUS_FLAGS_HEX_COLUMN)
    STATUS_FLAGS_COLUMN = 9
    worksheet.cell(row=3, column=STATUS_FLAGS_COLUMN).value = most_common_status_flags
    FREQUENCY_COLUMN = 10
    worksheet.cell(row=3, column=FREQUENCY_COLUMN).value = amount_of_status_flags
    AVERAGE_PERIOD_COLUMN = 11
    average_period = get_average_period(worksheet)
    worksheet.cell(row=3, column=AVERAGE_PERIOD_COLUMN).value = average_period
    sintese_worsheet = workbook['Sintese']
    export_data_into_summary_header(worksheet, sintese_worsheet, pmu)
    workbook.save(excel_file_name)


FIRST_LINE=1
FIRST_COLUMN=1
LAST_COLUMN=7
def header_in_excel(worksheet, pmu):
    worksheet.merge_cells(start_row=FIRST_LINE, end_row=FIRST_LINE, start_column=FIRST_LINE, end_column=LAST_COLUMN)
    worksheet.cell(row=1, column=1, value=pmu)
    worksheet.cell(row=2, column=1, value='Número')
    worksheet.cell(row=2, column=2, value='Data')
    worksheet.cell(row=2, column=3, value='Status Flags openPDC (hex)')
    worksheet.cell(row=2, column=4, value='Status Flags openPDC (binário)')
    worksheet.cell(row=2, column=5, value='Início (UTC)')
    worksheet.cell(row=2, column=6, value='Fim (UTC)')
    worksheet.cell(row=2, column=7, value='Período')
    for col in ['A', 'B', 'E', 'F', 'G']:
        worksheet.column_dimensions[col].auto_size = True
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 12
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
    BLUE_COLOR = 'FF0070C0'
    color_fill = PatternFill(start_color=BLUE_COLOR, end_color=BLUE_COLOR, fill_type='solid')  
    for cell in worksheet["1:1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = cell.font.copy(bold=True, size=12, name='Arial')
        cell.fill = color_fill
        if cell.value is not None:
            cell.value = cell.value.upper()
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

def configure_pmu_summary_header(worksheet):
    worksheet['I2'] = "Status Flag"
    worksheet['J2'] = "Frequência"
    worksheet['K2'] = "Período médio"
    worksheet['L2'] = "Período total"
    light_blue = 'ADD8E6'
    color_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type='solid')    
    for col in range(9, 13):
        cell = worksheet.cell(row=2, column=col)
        cell.fill = color_fill      
    for col in worksheet.iter_cols(min_col=9):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')


def get_period_time_of_status_flags(worksheet, status_flags, date):
    last_row = worksheet.max_row
    while worksheet.cell(row=last_row, column=2).value is None:
        last_row -= 1
    increment_by_two = range(0, len(status_flags), 2)
    for current_excel_line, item in enumerate(increment_by_two, start=last_row + 1):
        worksheet.insert_rows(current_excel_line)
        NUMBER_COLUMN = 1
        worksheet.cell(row=current_excel_line, column=NUMBER_COLUMN).value = current_excel_line - 2
        DATE_COLUMN = 2
        worksheet.cell(row=current_excel_line, column=DATE_COLUMN).value = datetime.datetime.strptime(
            date, '%m-%d-%y').strftime('%d/%m/%Y')
        STATUS_FLAGS_HEX_COLUMN = 3
        worksheet.cell(row=current_excel_line, column=STATUS_FLAGS_HEX_COLUMN).value = hex(
            status_flags[item]['Value'])[2:].upper()
        STATUS_FLAGS_BINARY_COLUMN = 4
        worksheet.cell(row=current_excel_line, column=STATUS_FLAGS_BINARY_COLUMN).value = convert_decimal_to_binary(
            status_flags)
        START_TIME_COLUMN = 5
        start_time = status_flags[item]['Time']
        worksheet.cell(row=current_excel_line, column=START_TIME_COLUMN).value = format_time(start_time)
        END_TIME_COLUMN = 6
        end_time = status_flags[item + 1]['Time']
        worksheet.cell(row=current_excel_line, column=END_TIME_COLUMN).value = format_time(end_time)
        PERIOD_COLUMN = 7
        duration_time = get_time_duration(start_time, end_time)
        worksheet.cell(row=current_excel_line, column=PERIOD_COLUMN).value = duration_time
        START_TIME_COLUMN = 5
        cell_start_time = worksheet.cell(row=current_excel_line, column=START_TIME_COLUMN)
        cell_start_time.number_format = "h:mm:ss.000"

def get_total_duration_of_status_flags(worksheet):
    PERIOD_COLUMN = 7
    total_duration = datetime.timedelta()
    for row in range(3, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=PERIOD_COLUMN).value
        if cell_value is not None:
            time_value = datetime.datetime.strptime(
                cell_value, "%H:%M:%S.%f")
            time_delta = datetime.timedelta(hours=time_value.hour, minutes=time_value.minute,
                                            seconds=time_value.second, microseconds=time_value.microsecond)
            total_duration += time_delta
    return total_duration

def get_total_items_in_an_excel_column(worksheet, column):
    status_flags = []
    count_total_status_flags = 0
    for row in worksheet.iter_rows(min_row=3, min_col=column, values_only=True):
        if row[0] is not None:
            status_flags.append(row[0])
            count_total_status_flags += 1
    counter = Counter(status_flags)
    most_common_status_flags, count = counter.most_common(1)[0]
    return most_common_status_flags, count_total_status_flags

TOTAL_PERIOD_COLUMN = 12
FREQUENCY_COLUMN = 10
def get_average_period(worksheet):
    average_period = worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value
    frequency = worksheet.cell(row=3, column=FREQUENCY_COLUMN).value
    return average_period / frequency

def export_data_into_summary_header(worksheet, sintese_worksheet, pmu):
    for row in sintese_worksheet.iter_rows():
        for cell in row:
            if cell.value == pmu:
                STATUS_FLAGS_COLUMN = 9
                if worksheet.cell(row=3, column=STATUS_FLAGS_COLUMN).value == 'DE1F0':
                    FREQUENCY_COLUMN = 10
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 1).value = worksheet.cell(row=3, column=FREQUENCY_COLUMN).value
                    AVERAGE_PERIOD_COLUMN = 11
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 2).value = worksheet.cell(row=3, column=AVERAGE_PERIOD_COLUMN).value
                    TOTAL_PERIOD_COLUMN = 12
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 3).value = worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value
                elif worksheet.cell(row=3, column=STATUS_FLAGS_COLUMN).value == 'DE040':
                    FREQUENCY_COLUMN = 10
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 4).value = worksheet.cell(row=3, column=FREQUENCY_COLUMN).value
                    AVERAGE_PERIOD_COLUMN = 11
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 5).value = worksheet.cell(row=3, column=AVERAGE_PERIOD_COLUMN).value
                    TOTAL_PERIOD_COLUMN = 12
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 6).value = worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value
                else:
                    FREQUENCY_COLUMN = 10
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 7).value = worksheet.cell(row=3, column=FREQUENCY_COLUMN).value
                    AVERAGE_PERIOD_COLUMN = 11
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 8).value = worksheet.cell(row=3, column=AVERAGE_PERIOD_COLUMN).value
                    TOTAL_PERIOD_COLUMN = 12
                    sintese_worksheet.cell(row=cell.row, column=cell.column + 9).value = worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value