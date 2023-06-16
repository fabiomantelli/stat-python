from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment

FIRST_LINE=1
FIRST_COLUMN=1
LAST_COLUMN=7
def header_in_excel(worksheet, pmu):
    worksheet.merge_cells(start_row=FIRST_LINE, end_row=FIRST_LINE, start_column=FIRST_LINE, end_column=LAST_COLUMN)
    worksheet.cell(row=1, column=1, value=pmu)

    worksheet.cell(row=2, column=1, value='Número')
    worksheet.cell(row=2, column=2, value='Data')
    worksheet.cell(row=2, column=3, value='Status Flags openPDC (hex)')
    worksheet.cell(row=2, column=4, value='Status Flags openPDC (binário)')
    worksheet.cell(row=2, column=5, value='Início (UTC)')
    worksheet.cell(row=2, column=6, value='Fim (UTC)')
    worksheet.cell(row=2, column=7, value='Período')

    for col in ['A', 'B', 'E', 'F', 'G']:
        worksheet.column_dimensions[col].auto_size = True
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 12
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15

    BLUE_COLOR = 'FF0070C0'
    color_fill = PatternFill(start_color=BLUE_COLOR, end_color=BLUE_COLOR, fill_type='solid')  

    for cell in worksheet["1:1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = cell.font.copy(bold=True, size=12, name='Arial')
        cell.fill = color_fill
        if cell.value is not None:
            cell.value = cell.value.upper()

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
