import openpyxl
from openpyxl import Workbook

from .header_in_excel import header_in_excel
from .configure_pmu_summary_header import configure_pmu_summary_header
from .get_total_duration_of_status_flags import get_total_duration_of_status_flags
from .get_period_time_of_status_flags import get_period_time_of_status_flags
from .get_average_period import get_average_period
from app.utils.get_total_items_in_a_excel_column import get_total_items_in_an_excel_column
from .export_data_into_summary_header import export_data_into_summary_header

def export_data_into_excel(status_flags, pmu, date, server_name):
    month, day, year = date.split("-")
    if not status_flags:
        return
    try:
        excel_file_name = f"./data/{year}_{month}_{server_name}_status_flags.xlsx"
        workbook = openpyxl.load_workbook(excel_file_name)
    except FileNotFoundError:
        workbook = Workbook()
    try:
        worksheet = workbook[pmu]
    except KeyError:
        worksheet = workbook.create_sheet(pmu)
    header_in_excel(worksheet, pmu)
    configure_pmu_summary_header(worksheet)
    get_period_time_of_status_flags(worksheet, status_flags, date)
    TOTAL_PERIOD_COLUMN = 12
    total_duration_of_status_flags = get_total_duration_of_status_flags(worksheet)
    worksheet.cell(row=3, column=TOTAL_PERIOD_COLUMN).value = total_duration_of_status_flags
    STATUS_FLAGS_HEX_COLUMN = 3
    most_common_status_flags, amount_of_status_flags  = get_total_items_in_an_excel_column(worksheet, STATUS_FLAGS_HEX_COLUMN)
    STATUS_FLAGS_COLUMN = 9
    worksheet.cell(row=3, column=STATUS_FLAGS_COLUMN).value = most_common_status_flags
    FREQUENCY_COLUMN = 10
    worksheet.cell(row=3, column=FREQUENCY_COLUMN).value = amount_of_status_flags
    AVERAGE_PERIOD_COLUMN = 11
    average_period = get_average_period(worksheet)
    worksheet.cell(row=3, column=AVERAGE_PERIOD_COLUMN).value = average_period
    sintese_worsheet = workbook['Síntese']
    export_data_into_summary_header(worksheet, sintese_worsheet, pmu)
    workbook.save(excel_file_name)
    

