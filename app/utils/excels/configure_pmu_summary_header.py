from openpyxl.styles import PatternFill 
from openpyxl.styles.alignment import Alignment

def configure_pmu_summary_header(worksheet):
    worksheet['I2'] = "Status Flag"
    worksheet['J2'] = "Frequência"
    worksheet['K2'] = "Período médio"
    worksheet['L2'] = "Período total"
    light_blue = 'ADD8E6'
    color_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type='solid')    
    for col in range(9, 13):
        cell = worksheet.cell(row=2, column=col)
        cell.fill = color_fill      
    for col in worksheet.iter_cols(min_col=9):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')