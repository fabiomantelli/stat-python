import datetime
import openpyxl

from openpyxl.styles.borders import Border, Side

from .excel_functions.configure_pmu_summary_header import configure_pmu_summary_header

def find_cell_value(sheet, target_value, col):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == col and cell.value == target_value:
                return cell
    return None

def formatting(pmu, date, server_name):
    month, day, year = date.split("-")
    excel_file_name = f"./data/{year}_{month}_medfasee_{server_name}_status_flags.xlsx"
    workbook = openpyxl.load_workbook(excel_file_name)  

    try:
        worksheet = workbook[pmu]
    except KeyError:
        return

    configure_pmu_summary_header(worksheet)
    
    groups = {}
    STATUS_FLAGS_HEX_COLUMN = 3
    for row in range(STATUS_FLAGS_HEX_COLUMN, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=STATUS_FLAGS_HEX_COLUMN).value

        if cell_value is not None:
            if cell_value not in groups:
                groups[cell_value] = []
            groups[cell_value].append(row)

    for group in groups.values():
        total_duration = 0

        for row in group:
            cell_value = worksheet.cell(row=row, column=7).value

            if cell_value is not None:
                duration_parts = cell_value.split(':')
                hours = int(duration_parts[0])
                minutes = int(duration_parts[1])
                seconds = int(duration_parts[2].split('.')[0])
                milliseconds = int(duration_parts[2].split('.')[1])
                total_duration += (hours * 3600) + (minutes * 60) + seconds + (milliseconds / 1000)

        td = datetime.timedelta(seconds=total_duration)

        formatted_duration = "{:02d}:{:02d}:{:02d}.{:03d}".format(td.seconds // 3600, (td.seconds // 60) % 60, td.seconds % 60, td.microseconds // 1000)

        TOTAL_PERIOD_COLUMN = 12
        worksheet.cell(row=group[0], column=TOTAL_PERIOD_COLUMN).value = formatted_duration

    element_durations = {}

    for row in range(3, worksheet.max_row + 1):
        element_value = worksheet.cell(row=row, column=3).value
        cell_value = worksheet.cell(row=row, column=7).value

        if cell_value is not None:
            duration_parts = cell_value.split(':')
            hours = int(duration_parts[0])
            minutes = int(duration_parts[1])
            seconds = int(duration_parts[2].split('.')[0])
            milliseconds = int(duration_parts[2].split('.')[1])
            total_duration = (hours * 3600) + (minutes * 60) + seconds + (milliseconds / 1000)

            if element_value not in element_durations:
                element_durations[element_value] = {"total_duration": total_duration, "count": 1}
            else:
                element_durations[element_value]["total_duration"] += total_duration
                element_durations[element_value]["count"] += 1

    for element, durations in element_durations.items():
        average_duration = durations["total_duration"] / durations["count"]
        td = datetime.timedelta(seconds=average_duration)
        formatted_duration = "{:02d}:{:02d}:{:02d}.{:03d}".format(td.seconds // 3600, (td.seconds // 60) % 60, td.seconds % 60, td.microseconds // 1000)
        worksheet.cell(row=3 + list(element_durations.keys()).index(element), column=11).value = formatted_duration

    valores_unicos = set()
    for row in worksheet.iter_rows(min_row=3, min_col=3, values_only=True):
        valores_unicos.add(row[0])

    valores_e_frequencias = {}
    for valor in valores_unicos:
        frequencia = 0
        for row in worksheet.iter_rows(min_row=3, min_col=3, values_only=True):
            if row[0] == valor:
                frequencia += 1
        valores_e_frequencias[valor] = frequencia

    cont = 0
    for row in worksheet.iter_rows(min_row=3):
        valor = row[2].value
        frequencia = valores_e_frequencias[valor]
        row[8].value = valor
        row[9].value = frequencia
        cont += 1
        if cont == len(valores_unicos):
            workbook.save(excel_file_name)
            break

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
    empty_border = Border()

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.border = empty_border
            
    worksheet = workbook[pmu]
    # Encontrando a célula "DE1F0" e obtendo sua posição
    de1f0_cell = find_cell_value(worksheet, "DE1F0", 9)
    if de1f0_cell is not None:
        de1f0_col = de1f0_cell.column
        de1f0_row = de1f0_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 1).value
        periodo_medio_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 2).value
        periodo_total_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 3).value
    else:
        freq_DE1F0 = 0
        periodo_medio_DE1F0 = 0
        periodo_total_DE1F0 = 0
    
    # Encontrando a célula "DE040" e obtendo sua posição
    de040_cell = find_cell_value(worksheet, "DE040", 9)
    if de040_cell is not None:
        de040_col = de040_cell.column
        de040_row = de040_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE040 = worksheet.cell(row=de040_row, column=de040_col + 1).value
        periodo_medio_DE040 = worksheet.cell(row=de040_row, column=de040_col + 2).value
        periodo_total_DE040 = worksheet.cell(row=de040_row, column=de040_col + 3).value
    else:
        freq_DE040 = 0
        periodo_medio_DE040 = 0
        periodo_total_DE040 = 0

    # Encontrando a célula "DE000" e obtendo sua posição
    de000_cell = find_cell_value(worksheet, "DE000", 9)
    if de000_cell is not None:
        de000_col = de000_cell.column
        de000_row = de000_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE000 = worksheet.cell(row=de000_row, column=de000_col + 1).value
        periodo_medio_DE000 = worksheet.cell(row=de000_row, column=de000_col + 2).value
        periodo_total_DE000 = worksheet.cell(row=de000_row, column=de000_col + 3).value
    else:
        freq_DE000 = 0
        periodo_medio_DE000 = 0
        periodo_total_DE000 = 0

    status_flags_diferente = 0
    column = worksheet['I']
    # Verificar se há valores diferentes de "DE1F0", "DE040" e "DE000" na coluna
    for cell in column:
        if cell.value and cell.value not in ['DE1F0', 'DE040', 'DE000','Status Flag']:
            status_flags_diferente = cell.value            
            break
    
    status_flags_diferente_cell = find_cell_value(worksheet, status_flags_diferente, 9)
    if  status_flags_diferente_cell is not None:
         status_flags_diferente_col = status_flags_diferente_cell.column
         status_flags_diferente_row = status_flags_diferente_cell.row
        # Obtendo os valores nas colunas adjacentes
         freq_status_flags_diferente = worksheet.cell(row=status_flags_diferente_row, column=status_flags_diferente_col + 1).value
         periodo_medio_status_flags_diferente = worksheet.cell(row=status_flags_diferente_row, column=status_flags_diferente_col + 2).value
         periodo_total_status_flags_diferente = worksheet.cell(row=status_flags_diferente_row, column=status_flags_diferente_col + 3).value
    else:
        freq_status_flags_diferente = 0
        periodo_medio_status_flags_diferente = 0
        periodo_total_status_flags_diferente = 0
     
    worksheet1 = workbook['Síntese']
    worksheet_exists = 'Síntese' in workbook.sheetnames
    
    if  worksheet_exists:         
        identificação = find_cell_value(worksheet1, pmu, 3)    
        identificação_col = identificação.column
        identificação_row = identificação.row       
        worksheet1.cell(row=identificação_row, column=identificação_col + 1).value = freq_DE1F0
        worksheet1.cell(row=identificação_row, column=identificação_col + 2).value = periodo_medio_DE1F0 
        worksheet1.cell(row=identificação_row, column=identificação_col + 3).value = periodo_total_DE1F0
        worksheet1.cell(row=identificação_row, column=identificação_col + 4).value = freq_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 5).value = periodo_medio_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 6).value = periodo_total_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 7).value = freq_DE000
        worksheet1.cell(row=identificação_row, column=identificação_col + 8).value = periodo_medio_DE000
        worksheet1.cell(row=identificação_row, column=identificação_col + 9).value = periodo_total_DE000 
        worksheet1.cell(row=identificação_row, column=identificação_col + 10).value = freq_status_flags_diferente
        worksheet1.cell(row=identificação_row, column=identificação_col + 11).value = periodo_medio_status_flags_diferente
        worksheet1.cell(row=identificação_row, column=identificação_col + 12).value = periodo_total_status_flags_diferente  
    
    workbook.save(excel_file_name)