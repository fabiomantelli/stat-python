import datetime
import openpyxl

from utils.create_file_name import create_file_name

def summary_pmu(pmu, date, server_name):
    file_name = create_file_name(date, server_name)  

    workbook = openpyxl.load_workbook(file_name)
    worksheet = get_worksheet(workbook, pmu)

    if not worksheet:
        return
        
    generate_header_for_summary(worksheet)
    
    groups = get_groups(worksheet)

    for group in groups.values():
        formatted_duration = calculate_group_duration(worksheet, group)
        worksheet.cell(row=group[0], column=12).value = formatted_duration

    element_durations = calculate_element_durations(worksheet)
    element_durations = calculate_average_duration(element_durations)

    for element, durations in element_durations.items():
        formatted_duration = durations["formatted_duration"]
        worksheet.cell(row=3 + list(element_durations.keys()).index(element), column=11).value = formatted_duration

    unique_values = get_unique_values(worksheet)
    values_and_frequencies = calculate_value_frequencies(worksheet, unique_values)
    write_value_frequencies(worksheet, values_and_frequencies)

    workbook.save(file_name)
       

def get_worksheet(workbook, pmu):
    if pmu in workbook.sheetnames:
        return workbook[pmu]
    else:
        return None

def generate_header_for_summary(worksheet):
    if worksheet:
        worksheet['I2'] = "Status Flag"
        worksheet['J2'] = "Frequência"
        worksheet['K2'] = "Período médio"
        worksheet['L2'] = "Período total"

def get_groups(worksheet):
    groups = {}
    for row in range(3, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=3).value
        if cell_value is not None:
            if cell_value not in groups:
                groups[cell_value] = []
            groups[cell_value].append(row)
    return groups

def calculate_group_duration(worksheet, group):
    total_duration = 0
    for row in group:
        cell_value = worksheet.cell(row=row, column=7).value
        if cell_value is not None:
            duration_parts = cell_value.split(':')
            hours = int(duration_parts[0])
            minutes = int(duration_parts[1])
            seconds = int(duration_parts[2].split('.')[0])
            milliseconds = int(duration_parts[2].split('.')[1])
            total_duration += (hours * 3600) + (minutes * 60) + seconds + (milliseconds / 1000)
    td = datetime.timedelta(seconds=total_duration)
    formatted_duration = "{:02d}:{:02d}:{:02d}.{:03d}".format(td.seconds // 3600, (td.seconds // 60) % 60, td.seconds % 60, td.microseconds // 1000)
    return formatted_duration

def calculate_element_durations(worksheet):
    element_durations = {}
    for row in range(3, worksheet.max_row + 1):
        element_value = worksheet.cell(row=row, column=3).value
        cell_value = worksheet.cell(row=row, column=7).value
        if cell_value is not None:
            duration_parts = cell_value.split(':')
            hours = int(duration_parts[0])
            minutes = int(duration_parts[1])
            seconds = int(duration_parts[2].split('.')[0])
            milliseconds = int(duration_parts[2].split('.')[1])
            total_duration = (hours * 3600) + (minutes * 60) + seconds + (milliseconds / 1000)
            if element_value not in element_durations:
                element_durations[element_value] = {"total_duration": total_duration, "count": 1}
            else:
                element_durations[element_value]["total_duration"] += total_duration
                element_durations[element_value]["count"] += 1
    return element_durations

def calculate_average_duration(element_durations):
    for element, durations in element_durations.items():
        average_duration = durations["total_duration"] / durations["count"]
        td = datetime.timedelta(seconds=average_duration)
        formatted_duration = "{:02d}:{:02d}:{:02d}.{:03d}".format(td.seconds // 3600, (td.seconds // 60) % 60, td.seconds % 60, td.microseconds // 1000)
        durations["formatted_duration"] = formatted_duration
    return element_durations

def get_unique_values(worksheet):
    unique_values = set()
    for row in worksheet.iter_rows(min_row=3, min_col=3, values_only=True):
        unique_values.add(row[0])
    return unique_values

def calculate_value_frequencies(worksheet, unique_values):
    values_and_frequencies = {}
    for valor in unique_values:
        frequencia = 0
        for row in worksheet.iter_rows(min_row=3, min_col=3, values_only=True):
            if row[0] == valor:
                frequencia += 1
        values_and_frequencies[valor] = frequencia
    return values_and_frequencies

def write_value_frequencies(worksheet, values_and_frequencies):
    cont = 0
    for row in worksheet.iter_rows(min_row=3):
        valor = row[2].value
        frequencia = values_and_frequencies[valor]
        row[8].value = valor
        row[9].value = frequencia
        cont += 1
        if cont == len(values_and_frequencies):
            break
   


