import openpyxl

from openpyxl.styles import Alignment, Border, Side, PatternFill
from utils.create_file_name import create_file_name

def format_workbook(pmu, date, server_name):

    file_name = create_file_name(date, server_name)
    workbook = openpyxl.load_workbook(file_name)
    worksheet_exists = pmu in workbook.sheetnames

    if worksheet_exists:
        worksheet = workbook[pmu]
    else:
        return

    adjust_column_sizes(worksheet)
    center_cell_content(worksheet)
    apply_borders(worksheet)
    remove_empty_cell_borders(worksheet)
    set_fill_colors(worksheet)
    
    workbook.save(file_name)

def adjust_column_sizes(worksheet):
    for col in ['A', 'B', 'E', 'F', 'G']:
        worksheet.column_dimensions[col].auto_size = True
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 12
    worksheet.column_dimensions['K'].width = 15
    worksheet.column_dimensions['L'].width = 15

def center_cell_content(worksheet):
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

def apply_borders(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

def remove_empty_cell_borders(worksheet):
    empty_border = Border()
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.border = empty_border

def set_fill_colors(worksheet):
    color_fill1 = PatternFill(start_color='FF0070C0', end_color='FF0070C0', fill_type='solid')
    color_fill2 = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    worksheet['A1'].fill = color_fill1
    worksheet['A1'].alignment = Alignment(horizontal='center')
    for col in range(9, 13):
        cell = worksheet.cell(row=2, column=col)
        cell.fill = color_fill2
