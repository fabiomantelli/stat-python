import datetime
import openpyxl
from openpyxl import Workbook

from utils.header_in_excel import header_in_excel


def export_data_into_excel(status_flags, pmu, date):
    parts = date.split("-")

    # Extract the day and month values from the parts list
    year = parts[2]
    month = parts[0]

    try:
        file_name = f"./data/{year}_{month}_medfasee_status_flags.xlsx"
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    if status_flags == []:
        return

    worksheet_exists = pmu in workbook.sheetnames

    if worksheet_exists:
        worksheet = workbook[pmu]
    else:
        worksheet = workbook.create_sheet(pmu)

    header_in_excel(worksheet, pmu)

    workbook.save(file_name)

    last_row = worksheet.max_row
    while worksheet.cell(row=last_row, column=2).value is None:
        last_row -= 1

    row_increment = 1

    for i in range(0, len(status_flags), 2):
        # Convert decimal to binary
        binary_string = bin(status_flags[i]['Value'])[2:]
        # zero-pad to a multiple of 8 bits
        binary_string = "0" * (8 - len(binary_string) % 8) + binary_string
        binary_list = [binary_string[i:i+4]
                       for i in range(0, len(binary_string), 4)]
        binary_with_spaces = " ".join(binary_list)

        new_line = last_row + row_increment
        worksheet.insert_rows(new_line)
        worksheet.cell(row=new_line, column=1).value = last_row-2+row_increment
        worksheet.cell(row=new_line, column=2).value = date
        worksheet.cell(row=new_line, column=3).value = hex(
            status_flags[i]['Value'])[2:].upper()
        worksheet.cell(row=new_line, column=4).value = binary_with_spaces

        # Convert time strings to datetime objects
        initial_time = datetime.datetime.strptime(
            status_flags[i]['Time'], "%Y-%m-%d %H:%M:%S.%f")
        final_time = datetime.datetime.strptime(
            status_flags[i+1]['Time'], "%Y-%m-%d %H:%M:%S.%f")

       # Calculate duration in seconds
        duration = (final_time - initial_time).total_seconds()

        worksheet.cell(row=new_line, column=5).value = initial_time.strftime(
            "%H:%M:%S.%f")[:-3]
        worksheet.cell(row=new_line, column=6).value = final_time.strftime(
            "%H:%M:%S.%f")[:-3]

        # Format duration as h:mm:ss.000
        hours, remainder = divmod(duration, 3600)
        minutes, seconds = divmod(remainder, 60)
        milliseconds = round((seconds - int(seconds)) * 1000)
        duration_str = "{:02d}:{:02d}:{:02d}.{:03d}".format(
            int(hours), int(minutes), int(seconds), milliseconds)
        worksheet.cell(row=new_line, column=7).value = duration_str

        # set cell format in Excel
        cell_initial_time = worksheet.cell(row=new_line, column=5)
        cell_initial_time.number_format = "h:mm:ss.000"
        cell_final_time = worksheet.cell(row=new_line, column=6)
        cell_final_time.number_format = "h:mm:ss.000"
        cell_period = worksheet.cell(row=new_line, column=7)
        cell_period.number_format = "h:mm:ss.000"
        row_increment += 1

        workbook.save(file_name)
