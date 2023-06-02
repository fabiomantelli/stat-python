import openpyxl
from utils.create_file_name import create_file_name

def update_síntese(pmu, date, server_name):

    file_name = create_file_name(date, server_name)
    workbook = openpyxl.load_workbook(file_name)
    worksheet = get_worksheet(workbook, pmu)

    if not worksheet:
        return

    freq_DE1F0, average_period_DE1F0, total_period_DE1F0 = get_cell_values(worksheet, "DE1F0")
    freq_DE040, average_period_DE040, total_period_DE040 = get_cell_values(worksheet, "DE040")
    freq_DE000, average_period_DE000, total_period_DE000 = get_cell_values(worksheet, "DE000")

    status_different_flags, frequency_status_different_flags, medium_period_different_status_flags, total_period_different_status_flags = get_status_flags_diff(worksheet)

    workbook.save(file_name)
    
    write_síntese(workbook, pmu, freq_DE1F0, average_period_DE1F0, total_period_DE1F0, freq_DE040,
                             average_period_DE040, total_period_DE040, freq_DE000,average_period_DE000, total_period_DE000, 
                             status_different_flags, frequency_status_different_flags, medium_period_different_status_flags, 
                             total_period_different_status_flags, server_name, date)


def get_worksheet(workbook, pmu):
    if pmu in workbook.sheetnames:
        return workbook[pmu]
    else:
        return None


def find_cell_value(sheet, target_value, col):    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == col and cell.value == target_value:
                return cell
    return None


def get_cell_values(worksheet, cell_value):
    cell = find_cell_value(worksheet, cell_value, 9)
    if cell is not None:
        col = cell.column
        row = cell.row
        freq = worksheet.cell(row=row, column=col + 1).value
        periodo_medio = worksheet.cell(row=row, column=col + 2).value
        periodo_total = worksheet.cell(row=row, column=col + 3).value
    else:
        freq = 0
        periodo_medio = 0
        periodo_total = 0

    return freq, periodo_medio, periodo_total


def get_status_flags_diff(worksheet):
    status_flags_diff = 0
    column = worksheet['I']    
    for cell in column:
        if cell.value and cell.value not in ['DE1F0', 'DE040', 'DE000', 'Status Flag']:
            status_flags_diff = cell.value
            break

    freq_status_flags_diff, average_period_status_flags_diff, total_period_status_flags_diff = 0, 0, 0

    status_flags_diff_cell = find_cell_value(worksheet, status_flags_diff, 9)
    if status_flags_diff_cell is not None:
        status_flags_diff_col = status_flags_diff_cell.column
        status_flags_diff_row = status_flags_diff_cell.row
        freq_status_flags_diff = worksheet.cell(row=status_flags_diff_row, column=status_flags_diff_col + 1).value
        average_period_status_flags_diff = worksheet.cell(row=status_flags_diff_row, column=status_flags_diff_col + 2).value
        total_period_status_flags_diff = worksheet.cell(row=status_flags_diff_row, column=status_flags_diff_col + 3).value

    return status_flags_diff, freq_status_flags_diff, average_period_status_flags_diff, total_period_status_flags_diff



def write_summary(workbook, pmu, freq_DE1F0, average_period_DE1F0, total_period_DE1F0, freq_DE040,
                  average_period_DE040, total_period_DE040, freq_DE000, average_period_DE000,
                  total_period_DE000, freq_status_flags_diff, average_period_status_flags_diff,
                  total_period_status_flags_diff, server_name, date):
    worksheet1 = workbook['Summary']
    worksheet_exists = 'Summary' in workbook.sheetnames

    if worksheet_exists:
        identification = find_cell_value(worksheet1, pmu, 3)
        identification_col = identification.column
        identification_row = identification.row
        worksheet1.cell(row=identification_row, column=identification_col + 1).value = freq_DE1F0
        worksheet1.cell(row=identification_row, column=identification_col + 2).value = average_period_DE1F0
        worksheet1.cell(row=identification_row, column=identification_col + 3).value = total_period_DE1F0
        worksheet1.cell(row=identification_row, column=identification_col + 4).value = freq_DE040
        worksheet1.cell(row=identification_row, column=identification_col + 5).value = average_period_DE040
        worksheet1.cell(row=identification_row, column=identification_col + 6).value = total_period_DE040
        worksheet1.cell(row=identification_row, column=identification_col + 7).value = freq_DE000
        worksheet1.cell(row=identification_row, column=identification_col + 8).value = average_period_DE000
        worksheet1.cell(row=identification_row, column=identification_col + 9).value = total_period_DE000
        worksheet1.cell(row=identification_row, column=identification_col + 10).value = freq_status_flags_diff
        worksheet1.cell(row=identification_row, column=identification_col + 11).value = average_period_status_flags_diff
        worksheet1.cell(row=identification_row, column=identification_col + 12).value = total_period_status_flags_diff

    workbook.save(create_file_name(date, server_name))

