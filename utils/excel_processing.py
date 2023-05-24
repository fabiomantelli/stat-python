import openpyxl
from openpyxl import Workbook

def find_cell_value(sheet, target_value, col):
    # Encontra a posição da célula com o valor desejado na coluna especificada
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == col and cell.value == target_value:
                return cell
    return None


def process_workbook(pmu, nome_arquivo):
    # Carregando o arquivo Excel
    workbook = openpyxl.load_workbook(nome_arquivo)
    worksheet = workbook[pmu]
    # Encontrando a célula "DE1F0" e obtendo sua posição
    de1f0_cell = find_cell_value(worksheet, "DE1F0", 9)
    if de1f0_cell is not None:
        de1f0_col = de1f0_cell.column
        de1f0_row = de1f0_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 1).value
        periodo_medio_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 2).value
        periodo_total_DE1F0 = worksheet.cell(row=de1f0_row, column=de1f0_col + 3).value
    else:
        freq_DE1F0 = None
        periodo_medio_DE1F0 = None
        periodo_total_DE1F0 = None
    
    # Encontrando a célula "DE040" e obtendo sua posição
    de040_cell = find_cell_value(worksheet, "DE040", 9)
    if de040_cell is not None:
        de040_col = de040_cell.column
        de040_row = de040_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE040 = worksheet.cell(row=de040_row, column=de040_col + 1).value
        periodo_medio_DE040 = worksheet.cell(row=de040_row, column=de040_col + 2).value
        periodo_total_DE040 = worksheet.cell(row=de040_row, column=de040_col + 3).value
    else:
        freq_DE040 = None
        periodo_medio_DE040 = None
        periodo_total_DE040 = None

    
    # Encontrando a célula "DE000" e obtendo sua posição
    de000_cell = find_cell_value(worksheet, "DE000", 9)
    if de000_cell is not None:
        de000_col = de000_cell.column
        de000_row = de000_cell.row
        # Obtendo os valores nas colunas adjacentes
        freq_DE000 = worksheet.cell(row=de000_row, column=de000_col + 1).value
        periodo_medio_DE000 = worksheet.cell(row=de000_row, column=de000_col + 2).value
        periodo_total_DE000 = worksheet.cell(row=de000_row, column=de000_col + 3).value
    else:
        freq_DE000 = None
        periodo_medio_DE000 = None
        periodo_total_DE000 = None

    workbook.save(nome_arquivo)

    
    worksheet1 = workbook['Resumo']
    worksheet_exists = 'Resumo' in workbook.sheetnames
    
    if  worksheet_exists:         

        identificação = find_cell_value(worksheet1, pmu, 3)    
        identificação_col = identificação.column
        identificação_row = identificação.row       
        worksheet1.cell(row=identificação_row, column=identificação_col + 1).value = freq_DE1F0
        worksheet1.cell(row=identificação_row, column=identificação_col + 2).value = periodo_medio_DE1F0 
        worksheet1.cell(row=identificação_row, column=identificação_col + 3).value = periodo_total_DE1F0
        worksheet1.cell(row=identificação_row, column=identificação_col + 4).value = freq_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 5).value = periodo_medio_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 6).value = periodo_total_DE040
        worksheet1.cell(row=identificação_row, column=identificação_col + 7).value = freq_DE000
        worksheet1.cell(row=identificação_row, column=identificação_col + 8).value = periodo_medio_DE000
        worksheet1.cell(row=identificação_row, column=identificação_col + 9).value = periodo_total_DE000    
    print (worksheet1.cell(row=identificação_row, column=identificação_col + 1).value) 

    workbook.save(nome_arquivo)
    